"""V0.03 - Betreuungsplanung Niklas Tom-Hardy.

Änderungen ggü. V0.02:
- Zeitraster erweitert auf 07:30 - 18:00 (für 9h-Tage).
- Neue Tagesbausteine: office_long_day (9h), ho_long_day (9h).
- TM-Logik refactored: pro Tag freies Zeitfenster (nicht mehr fix 07:30-13:00).
- Neue Varianten:
    V10 - Paul 35h auf 4 Tage (2 HO + 2 Office, Fr frei) + Dom 20h (Mo/Di 8h + Do 4h)
          + Tagesmutter 20h (Mo 8h + Di 8h + Do 4h).
    V11 - Wie V10, aber Tagesmutter nur Mo/Di = 16h.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = r"C:\Users\Paul Elias\Desktop\PRU\Tom-Hardy Betreuung\Betreuungsplanung_Niklas_V0.04.xlsx"

# ---------- Farben / Stile ----------
COL_OFFICE     = PatternFill("solid", start_color="4F81BD")
COL_HO         = PatternFill("solid", start_color="9BBB59")
COL_PENDELN    = PatternFill("solid", start_color="BFBFBF")
COL_PAUSE_O    = PatternFill("solid", start_color="F2DCDB")
COL_PAUSE_H    = PatternFill("solid", start_color="F8CBAD")
COL_TM         = PatternFill("solid", start_color="B284BE")
COL_FREI       = PatternFill("solid", start_color="FFFFFF")
COL_HEADER     = PatternFill("solid", start_color="1F4E78")
COL_SUBHEAD    = PatternFill("solid", start_color="D9E1F2")
COL_TOTAL      = PatternFill("solid", start_color="FFF2CC")
COL_INCOME     = PatternFill("solid", start_color="E2EFDA")
COL_COST       = PatternFill("solid", start_color="FCE4D6")
COL_NIKLAS_OK  = PatternFill("solid", start_color="92D050")
COL_NIKLAS_GAP = PatternFill("solid", start_color="FF6B6B")

THIN = Side(border_style="thin", color="808080")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FONT_HEADER = Font(name="Arial", size=14, bold=True, color="FFFFFF")
FONT_SUBHEAD = Font(name="Arial", size=11, bold=True, color="1F4E78")
FONT_NORMAL = Font(name="Arial", size=10)
FONT_BOLD = Font(name="Arial", size=10, bold=True)
FONT_SMALL = Font(name="Arial", size=9)
FONT_LEGEND = Font(name="Arial", size=9, italic=True)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

# ---------- Zeit-Slots 07:30 - 17:30 (letzter Slot = 17:30-18:00) ----------
SLOTS = []
_h, _m = 7, 30
while (_h, _m) <= (17, 30):
    SLOTS.append(f"{_h:02d}:{_m:02d}")
    _m += 30
    if _m == 60:
        _m = 0
        _h += 1
DAYS = ["Montag", "Dienstag", "Mittwoch", "Donnerstag", "Freitag"]

# ---------- Status: (label, fill, location) ----------
S_OFF       = ("Büro",        COL_OFFICE,  "away")
S_HO        = ("HO",          COL_HO,      "home")
S_PEND_TO   = ("→Büro",       COL_PENDELN, "away")
S_PEND_HOME = ("→Home",       COL_PENDELN, "away")
S_PAU_OFF   = ("Pause",       COL_PAUSE_O, "away")
S_PAU_HOME  = ("Pause",       COL_PAUSE_H, "home")
S_TM        = ("Tagesmutter", COL_TM,      "tm")
S_FREI      = ("Frei",        COL_FREI,    "home")


def slot_idx(t):
    return SLOTS.index(t)


def fill_range(plan, day, t_start, t_end, status):
    i1 = SLOTS.index(t_start)
    if t_end in SLOTS:
        i2 = SLOTS.index(t_end)
    else:
        # Erlaubt z.B. "18:00" als Endpunkt (1 Slot nach SLOTS[-1])
        i2 = len(SLOTS)
    for i in range(i1, i2):
        plan[day][i] = status


def empty_plan():
    return {d: [S_FREI] * len(SLOTS) for d in DAYS}


# ---------- TM-Kosten Anlage 1a (gültig 01.01.2026 – 31.08.2026, 'unter 3 J.') ----------
TM_KOSTEN_U3 = {
    1: [0, 21, 42, 63, 85, 106, 127],
    2: [0, 35, 71, 106, 141, 176, 212],
    3: [0, 49, 99, 148, 197, 247, 296],
    4: [0, 63, 127, 190, 254, 317, 381],
    5: [0, 78, 155, 233, 310, 388, 465],
    6: [0, 92, 183, 275, 367, 458, 550],
    7: [0, 106, 212, 317, 423, 529, 635],
    8: [0, 113, 226, 338, 451, 564, 677],
}
GRUPPEN_LIMITS = [(1, "I", 20000), (2, "II", 40000), (3, "III", 60000),
                  (4, "IV", 80000), (5, "V", 100000), (6, "VI", 120000),
                  (7, "VII", float("inf"))]


def hours_to_bracket(h):
    if h < 5: return None
    if h <= 9: return 1
    if h <= 14: return 2
    if h <= 19: return 3
    if h <= 24: return 4
    if h <= 29: return 5
    if h <= 34: return 6
    if h <= 39: return 7
    return 8


def income_to_group(brutto_jahr):
    haushalt = brutto_jahr * 0.7
    for idx, name, limit in GRUPPEN_LIMITS:
        if haushalt <= limit:
            return idx, name, haushalt
    return 7, "VII", haushalt


def tm_kosten(brutto_jahr_haushalt, stunden_pro_woche):
    bracket = hours_to_bracket(stunden_pro_woche)
    if bracket is None:
        return 0, "—", 0
    grp_idx, grp_name, haushalt = income_to_group(brutto_jahr_haushalt)
    return TM_KOSTEN_U3[bracket][grp_idx - 1], grp_name, haushalt


def tm_hours_from_plan(tm_plan):
    return sum(0.5 for d in DAYS for s in tm_plan[d] if s == S_TM)


def make_standard_tm_plan(days_list):
    tm = empty_plan()
    for d in days_list:
        fill_range(tm, d, "07:30", "13:00", S_TM)
    return tm


# ---------- Niklas-Coverage ----------
def _label(s):
    if s == S_HO: return "HO"
    if s == S_FREI: return "frei"
    if s == S_PAU_HOME: return "Pause"
    return "?"


def derive_niklas(paul_plan, dom_plan, tm_plan=None):
    tm_plan = tm_plan or empty_plan()
    niklas = {}
    for d in DAYS:
        row = []
        for i, t in enumerate(SLOTS):
            tm = tm_plan[d][i]
            p = paul_plan[d][i]
            do = dom_plan[d][i]
            if tm == S_TM:
                row.append(("TM", COL_NIKLAS_OK))
                continue
            p_home = p[2] == "home"
            d_home = do[2] == "home"
            if p_home and d_home:
                if p == S_FREI and do == S_FREI:
                    row.append(("Beide frei", COL_NIKLAS_OK))
                else:
                    row.append((f"P:{_label(p)} D:{_label(do)}", COL_NIKLAS_OK))
            elif p_home:
                row.append((f"Paul {_label(p)}", COL_NIKLAS_OK))
            elif d_home:
                row.append((f"Dom {_label(do)}", COL_NIKLAS_OK))
            else:
                row.append(("LÜCKE!", COL_NIKLAS_GAP))
        niklas[d] = row
    return niklas


# ---------- Workbook ----------
wb = Workbook()
wb.remove(wb.active)


def write_legend(ws, start_row):
    ws.cell(row=start_row, column=1, value="Legende:").font = FONT_BOLD
    items = [
        ("Büro", COL_OFFICE),
        ("Homeoffice (HO) – ca. 50% Eigen-Betreuung möglich", COL_HO),
        ("Pendeln (→Büro / →Home)", COL_PENDELN),
        ("Pause @ Büro (unbezahlt)", COL_PAUSE_O),
        ("Pause @ Zuhause (unbezahlt)", COL_PAUSE_H),
        ("Tagesmutter", COL_TM),
        ("Frei / Eltern-Betreuung", COL_FREI),
        ("Niklas: betreut", COL_NIKLAS_OK),
        ("Niklas: LÜCKE", COL_NIKLAS_GAP),
    ]
    for i, (text, fill) in enumerate(items):
        c1 = ws.cell(row=start_row + 1 + i, column=1, value="")
        c1.fill = fill
        c1.border = BORDER
        c2 = ws.cell(row=start_row + 1 + i, column=2, value=text)
        c2.font = FONT_LEGEND
        ws.merge_cells(start_row=start_row + 1 + i, start_column=2,
                       end_row=start_row + 1 + i, end_column=5)
    return start_row + 1 + len(items)


def add_schedule(ws, start_row, title, paul_plan, dom_plan, niklas_plan):
    ws.cell(row=start_row, column=1, value=title).font = FONT_SUBHEAD
    ws.merge_cells(start_row=start_row, start_column=1,
                   end_row=start_row, end_column=16)
    ws.cell(row=start_row, column=1).fill = COL_SUBHEAD
    ws.cell(row=start_row, column=1).alignment = LEFT

    c = ws.cell(row=start_row + 1, column=1, value="Zeit")
    c.font = FONT_HEADER; c.fill = COL_HEADER; c.alignment = CENTER; c.border = BORDER
    ws.cell(row=start_row + 2, column=1, value="").border = BORDER
    ws.cell(row=start_row + 2, column=1).fill = COL_HEADER

    col = 2
    for d in DAYS:
        ws.merge_cells(start_row=start_row + 1, start_column=col,
                       end_row=start_row + 1, end_column=col + 2)
        c = ws.cell(row=start_row + 1, column=col, value=d)
        c.font = FONT_HEADER; c.fill = COL_HEADER; c.alignment = CENTER; c.border = BORDER
        col += 3

    col = 2
    for d in DAYS:
        for s in ["Paul", "Dom", "Niklas"]:
            c = ws.cell(row=start_row + 2, column=col, value=s)
            c.font = FONT_BOLD; c.fill = COL_SUBHEAD; c.alignment = CENTER; c.border = BORDER
            col += 1

    for i, t in enumerate(SLOTS):
        r = start_row + 3 + i
        c = ws.cell(row=r, column=1, value=t)
        c.font = FONT_SMALL; c.alignment = CENTER; c.border = BORDER; c.fill = COL_SUBHEAD
        col = 2
        for d in DAYS:
            for plan_item in (paul_plan[d][i], dom_plan[d][i], niklas_plan[d][i]):
                label, fill = plan_item[0], plan_item[1]
                cc = ws.cell(row=r, column=col, value=label)
                cc.fill = fill; cc.alignment = CENTER; cc.font = FONT_SMALL; cc.border = BORDER
                col += 1
    return start_row + 2 + len(SLOTS)


def _netto_formula(brutto_cell):
    """Approx Netto/Monat, Steuerklasse IV/IV verheiratet, 1 Kind, 2026."""
    return (f"=IF({brutto_cell}>=8000,{brutto_cell}*0.6,"
            f"IF({brutto_cell}>=7000,{brutto_cell}*0.61,"
            f"IF({brutto_cell}>=6000,{brutto_cell}*0.62,"
            f"IF({brutto_cell}>=5000,{brutto_cell}*0.64,"
            f"IF({brutto_cell}>=4000,{brutto_cell}*0.66,"
            f"IF({brutto_cell}>=3000,{brutto_cell}*0.68,"
            f"{brutto_cell}*0.70))))))")


def add_income_table(ws, start_row, paul_h, dom_h, paul_jahr=96000, dom_jahr=84000,
                     tm_kosten_monat=0, tm_label=""):
    r = start_row
    ws.cell(row=r, column=1, value="Einkommen & Kosten").font = FONT_SUBHEAD
    ws.cell(row=r, column=1).fill = COL_SUBHEAD
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    r += 1

    headers = ["", "Wochenstd.", "Vollzeit Brutto/J",
               "Anteil", "Brutto/Mon", "Brutto/Jahr",
               "Netto/Mon (≈)", "Netto/Jahr (≈)"]
    for i, h in enumerate(headers):
        c = ws.cell(row=r, column=1 + i, value=h)
        c.font = FONT_BOLD; c.fill = COL_TOTAL; c.alignment = CENTER; c.border = BORDER
    r += 1

    def write_person(name, hours, vollzeit):
        nonlocal r
        ws.cell(row=r, column=1, value=name).font = FONT_BOLD
        ws.cell(row=r, column=2, value=hours)
        ws.cell(row=r, column=3, value=vollzeit)
        ws.cell(row=r, column=4, value=f"=B{r}/40")
        ws.cell(row=r, column=5, value=f"=C{r}*D{r}/12")
        ws.cell(row=r, column=6, value=f"=C{r}*D{r}")
        ws.cell(row=r, column=7, value=_netto_formula(f"E{r}"))
        ws.cell(row=r, column=8, value=f"=G{r}*12")
        for col in range(1, 9):
            ws.cell(row=r, column=col).border = BORDER
            ws.cell(row=r, column=col).alignment = CENTER
        ws.cell(row=r, column=3).number_format = '#,##0 "€"'
        ws.cell(row=r, column=4).number_format = '0%'
        for col in (5, 6, 7, 8):
            ws.cell(row=r, column=col).number_format = '#,##0 "€"'
        result = r
        r += 1
        return result

    paul_row = write_person("Paul", paul_h, paul_jahr)
    dom_row = write_person("Dominique", dom_h, dom_jahr)

    ws.cell(row=r, column=1, value="Haushalt gesamt").font = FONT_BOLD
    ws.cell(row=r, column=2, value=f"=B{paul_row}+B{dom_row}")
    ws.cell(row=r, column=5, value=f"=E{paul_row}+E{dom_row}")
    ws.cell(row=r, column=6, value=f"=F{paul_row}+F{dom_row}")
    ws.cell(row=r, column=7, value=f"=G{paul_row}+G{dom_row}")
    ws.cell(row=r, column=8, value=f"=H{paul_row}+H{dom_row}")
    for col in range(1, 9):
        ws.cell(row=r, column=col).fill = COL_INCOME
        ws.cell(row=r, column=col).border = BORDER
        ws.cell(row=r, column=col).alignment = CENTER
        ws.cell(row=r, column=col).font = FONT_BOLD
    ws.cell(row=r, column=2).number_format = '0.0'
    for col in (5, 6, 7, 8):
        ws.cell(row=r, column=col).number_format = '#,##0 "€"'
    sum_row = r
    r += 2

    ws.cell(row=r, column=1, value="Haushaltseinkommen lt. Satzung (Brutto -30%)").font = FONT_BOLD
    ws.cell(row=r, column=2, value=f"=F{sum_row}*0.7")
    ws.cell(row=r, column=2).number_format = '#,##0 "€"'
    ws.cell(row=r, column=2).fill = COL_INCOME
    ws.cell(row=r, column=2).border = BORDER
    r += 1

    if tm_kosten_monat > 0:
        ws.cell(row=r, column=1, value=f"Tagesmutter-Kosten/Monat ({tm_label})").font = FONT_BOLD
        ws.cell(row=r, column=2, value=tm_kosten_monat)
        ws.cell(row=r, column=2).number_format = '#,##0 "€"'
        ws.cell(row=r, column=2).fill = COL_COST
        ws.cell(row=r, column=2).border = BORDER
        tm_row = r
        r += 1

        ws.cell(row=r, column=1, value="Netto-Haushalt nach TM-Kosten / Monat").font = FONT_BOLD
        ws.cell(row=r, column=2, value=f"=G{sum_row}-B{tm_row}")
        ws.cell(row=r, column=2).number_format = '#,##0 "€"'
        ws.cell(row=r, column=2).fill = COL_INCOME
        ws.cell(row=r, column=2).border = BORDER
        ws.cell(row=r, column=2).font = FONT_BOLD
        r += 1
    else:
        ws.cell(row=r, column=1, value="Netto-Haushalt / Monat (ohne TM)").font = FONT_BOLD
        ws.cell(row=r, column=2, value=f"=G{sum_row}")
        ws.cell(row=r, column=2).number_format = '#,##0 "€"'
        ws.cell(row=r, column=2).fill = COL_INCOME
        ws.cell(row=r, column=2).border = BORDER
        ws.cell(row=r, column=2).font = FONT_BOLD
        r += 1

    r += 1
    ws.cell(row=r, column=1,
            value="Netto-Schätzung: Steuerklasse IV/IV, verheiratet, 1 Kind, "
                  "2026 (Quelle: brutto-netto-rechner.info Näherung).").font = FONT_LEGEND
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    return r + 1


def setup_columns(ws):
    ws.column_dimensions['A'].width = 8
    for i in range(2, 17):
        ws.column_dimensions[get_column_letter(i)].width = 11


# ---------- Tagesbausteine ----------
def office_full_day(plan, day):
    """8h Büro."""
    fill_range(plan, day, "07:30", "08:00", S_PEND_TO)
    fill_range(plan, day, "08:00", "12:00", S_OFF)
    fill_range(plan, day, "12:00", "12:30", S_PAU_OFF)
    fill_range(plan, day, "12:30", "16:30", S_OFF)
    fill_range(plan, day, "16:30", "17:00", S_PEND_HOME)


def office_long_day(plan, day, hours=9):
    """9h Büro: 08-12 + Pause + 12:30-17:30."""
    if hours == 9:
        fill_range(plan, day, "07:30", "08:00", S_PEND_TO)
        fill_range(plan, day, "08:00", "12:00", S_OFF)
        fill_range(plan, day, "12:00", "12:30", S_PAU_OFF)
        fill_range(plan, day, "12:30", "17:30", S_OFF)
        fill_range(plan, day, "17:30", "18:00", S_PEND_HOME)


def ho_full_day(plan, day):
    """8h HO."""
    fill_range(plan, day, "08:00", "12:00", S_HO)
    fill_range(plan, day, "12:00", "12:30", S_PAU_HOME)
    fill_range(plan, day, "12:30", "16:30", S_HO)


def ho_long_day(plan, day, hours=9):
    """9h HO: 08-12 + Pause + 12:30-17:30."""
    if hours == 9:
        fill_range(plan, day, "08:00", "12:00", S_HO)
        fill_range(plan, day, "12:00", "12:30", S_PAU_HOME)
        fill_range(plan, day, "12:30", "17:30", S_HO)


def hybrid_day(plan, day, morning="office"):
    """Hybrid 8h."""
    if morning == "office":
        fill_range(plan, day, "07:30", "08:00", S_PEND_TO)
        fill_range(plan, day, "08:00", "12:00", S_OFF)
        fill_range(plan, day, "12:00", "12:30", S_PEND_HOME)
        fill_range(plan, day, "12:30", "13:00", S_PAU_HOME)
        fill_range(plan, day, "13:00", "17:00", S_HO)
    else:
        fill_range(plan, day, "08:00", "12:00", S_HO)
        fill_range(plan, day, "12:00", "12:30", S_PAU_HOME)
        fill_range(plan, day, "12:30", "13:00", S_PEND_TO)
        fill_range(plan, day, "13:00", "17:00", S_OFF)


def office_short_day(plan, day, hours):
    if hours == 7:
        fill_range(plan, day, "07:30", "08:00", S_PEND_TO)
        fill_range(plan, day, "08:00", "12:00", S_OFF)
        fill_range(plan, day, "12:00", "12:30", S_PAU_OFF)
        fill_range(plan, day, "12:30", "15:30", S_OFF)
        fill_range(plan, day, "15:30", "16:00", S_PEND_HOME)
    elif hours == 6:
        fill_range(plan, day, "07:30", "08:00", S_PEND_TO)
        fill_range(plan, day, "08:00", "14:00", S_OFF)
        fill_range(plan, day, "14:00", "14:30", S_PEND_HOME)
    elif hours == 4:
        fill_range(plan, day, "07:30", "08:00", S_PEND_TO)
        fill_range(plan, day, "08:00", "12:00", S_OFF)
        fill_range(plan, day, "12:00", "12:30", S_PEND_HOME)
    elif hours == 2:
        fill_range(plan, day, "07:30", "08:00", S_PEND_TO)
        fill_range(plan, day, "08:00", "10:00", S_OFF)
        fill_range(plan, day, "10:00", "10:30", S_PEND_HOME)


def ho_short_day(plan, day, hours):
    if hours == 7:
        fill_range(plan, day, "08:00", "12:00", S_HO)
        fill_range(plan, day, "12:00", "12:30", S_PAU_HOME)
        fill_range(plan, day, "12:30", "15:30", S_HO)
    elif hours == 6:
        fill_range(plan, day, "08:00", "14:00", S_HO)
    elif hours == 4:
        fill_range(plan, day, "08:00", "12:00", S_HO)
    elif hours == 2:
        fill_range(plan, day, "08:00", "10:00", S_HO)


def hybrid_short_day(plan, day, h_office, h_ho, morning="office"):
    """Hybrid mit reduzierten Stunden."""
    if morning == "office":
        fill_range(plan, day, "07:30", "08:00", S_PEND_TO)
        fill_range(plan, day, "08:00", "12:00", S_OFF)
        fill_range(plan, day, "12:00", "12:30", S_PEND_HOME)
        fill_range(plan, day, "12:30", "13:00", S_PAU_HOME)
        end_t = f"{13 + h_ho:02d}:00"
        fill_range(plan, day, "13:00", end_t, S_HO)
    else:
        fill_range(plan, day, "08:00", "12:00", S_HO)
        fill_range(plan, day, "12:00", "12:30", S_PAU_HOME)
        fill_range(plan, day, "12:30", "13:00", S_PEND_TO)
        end_t = f"{13 + h_office:02d}:00"
        fill_range(plan, day, "13:00", end_t, S_OFF)


# ---------- Varianten ----------

def make_v1():
    p, d = empty_plan(), empty_plan()
    office_full_day(p, "Montag");      ho_full_day(d, "Montag")
    office_full_day(p, "Dienstag");    ho_full_day(d, "Dienstag")
    hybrid_day(p, "Mittwoch", "office"); hybrid_day(d, "Mittwoch", "ho")
    ho_full_day(p, "Donnerstag");      office_full_day(d, "Donnerstag")
    ho_full_day(p, "Freitag");         office_full_day(d, "Freitag")
    return p, d


def make_v2():
    p, d = empty_plan(), empty_plan()
    office_full_day(p, "Montag");      ho_short_day(d, "Montag", 4)
    office_full_day(p, "Dienstag");    ho_short_day(d, "Dienstag", 4)
    hybrid_day(p, "Mittwoch", "office"); ho_short_day(d, "Mittwoch", 4)
    ho_full_day(p, "Donnerstag");      office_short_day(d, "Donnerstag", 4)
    ho_full_day(p, "Freitag");         office_short_day(d, "Freitag", 4)
    return p, d


def make_v3():
    p, d = empty_plan(), empty_plan()
    office_short_day(p, "Montag", 7);    ho_short_day(d, "Montag", 7)
    office_short_day(p, "Dienstag", 7);  ho_short_day(d, "Dienstag", 7)
    hybrid_short_day(p, "Mittwoch", 4, 3, "office")
    hybrid_short_day(d, "Mittwoch", 3, 4, "ho")
    ho_short_day(p, "Donnerstag", 7);    office_short_day(d, "Donnerstag", 7)
    ho_short_day(p, "Freitag", 7);       office_short_day(d, "Freitag", 7)
    return p, d


def make_v4():
    p, d = empty_plan(), empty_plan()
    office_short_day(p, "Montag", 7);      ho_full_day(d, "Montag")
    office_short_day(p, "Dienstag", 7);    ho_full_day(d, "Dienstag")
    hybrid_short_day(p, "Mittwoch", 4, 3, "office")
    hybrid_day(d, "Mittwoch", "ho")
    ho_short_day(p, "Donnerstag", 7);      office_full_day(d, "Donnerstag")
    office_full_day(d, "Freitag")
    return p, d


def make_v5():
    p, d = empty_plan(), empty_plan()
    office_short_day(p, "Montag", 7);    ho_full_day(d, "Montag")
    office_short_day(p, "Dienstag", 7);  ho_full_day(d, "Dienstag")
    ho_short_day(p, "Mittwoch", 7);      hybrid_day(d, "Mittwoch", "ho")
    office_full_day(d, "Donnerstag")
    office_full_day(d, "Freitag")
    return p, d


def make_v6():
    return make_v1()


def make_v7():
    return make_v3()


def make_v8():
    p, d = empty_plan(), empty_plan()
    office_short_day(p, "Montag", 7)
    office_short_day(p, "Mittwoch", 7)
    hybrid_short_day(p, "Donnerstag", 4, 3, "office")
    ho_short_day(p, "Freitag", 7)
    ho_short_day(d, "Montag", 4)
    ho_short_day(d, "Dienstag", 4)
    office_short_day(d, "Mittwoch", 4)
    ho_short_day(d, "Donnerstag", 4)
    office_short_day(d, "Freitag", 4)
    return p, d


def make_v9():
    p, d = empty_plan(), empty_plan()
    office_short_day(p, "Montag", 7)
    office_short_day(p, "Mittwoch", 7)
    ho_short_day(p, "Freitag", 7)
    office_short_day(d, "Montag", 4)
    ho_short_day(d, "Dienstag", 4)
    office_short_day(d, "Mittwoch", 4)
    ho_short_day(d, "Donnerstag", 4)
    ho_short_day(d, "Freitag", 4)
    return p, d


def make_v10_v11_paul_dom():
    """Paul 35h auf 4 Tage: Mo HO9, Di HO9, Mi Office9, Do Office8, Fr frei.
    Dom 20h: Mo Office8, Di Office8, Mi frei, Do HO4, Fr frei."""
    p, d = empty_plan(), empty_plan()
    ho_long_day(p, "Montag", 9)
    ho_long_day(p, "Dienstag", 9)
    office_long_day(p, "Mittwoch", 9)
    office_full_day(p, "Donnerstag")
    # Freitag: Paul frei
    office_full_day(d, "Montag")
    office_full_day(d, "Dienstag")
    # Mittwoch: Dom frei
    ho_short_day(d, "Donnerstag", 4)
    # Freitag: Dom frei
    return p, d


def make_v10_tm():
    """V10 TM: Mo 07:30-15:30 (8h), Di 07:30-15:30 (8h), Do 07:30-11:30 (4h) = 20h."""
    tm = empty_plan()
    fill_range(tm, "Montag", "07:30", "15:30", S_TM)
    fill_range(tm, "Dienstag", "07:30", "15:30", S_TM)
    fill_range(tm, "Donnerstag", "07:30", "11:30", S_TM)
    return tm


def make_v11_tm():
    """V11 TM: Mo 07:30-15:30 (8h), Di 07:30-15:30 (8h) = 16h."""
    tm = empty_plan()
    fill_range(tm, "Montag", "07:30", "15:30", S_TM)
    fill_range(tm, "Dienstag", "07:30", "15:30", S_TM)
    return tm


def make_v12():
    """V12: Paul wie V10. Dom Mo Office8, Di Office8, Mi frei, Do HO 8h, Fr frei = 24h."""
    p, d = empty_plan(), empty_plan()
    ho_long_day(p, "Montag", 9)
    ho_long_day(p, "Dienstag", 9)
    office_long_day(p, "Mittwoch", 9)
    office_full_day(p, "Donnerstag")
    office_full_day(d, "Montag")
    office_full_day(d, "Dienstag")
    ho_full_day(d, "Donnerstag")
    return p, d


def make_v12_tm():
    """V12 TM: Mo + Di + Do je 07:30-15:30 (8h) = 24h."""
    tm = empty_plan()
    fill_range(tm, "Montag", "07:30", "15:30", S_TM)
    fill_range(tm, "Dienstag", "07:30", "15:30", S_TM)
    fill_range(tm, "Donnerstag", "07:30", "15:30", S_TM)
    return tm


# ---------- Sheet-Builder ----------
def build_variant_sheet(name, title, subtitle, paul, dom, paul_h, dom_h,
                        tm_plan=None, tm_label="", tm_days_count=0):
    ws = wb.create_sheet(name)
    setup_columns(ws)
    ws.cell(row=1, column=1, value=title).font = FONT_HEADER
    ws.cell(row=1, column=1).fill = COL_HEADER
    ws.merge_cells("A1:P1")
    ws.cell(row=1, column=1).alignment = CENTER
    ws.cell(row=2, column=1, value=subtitle).font = FONT_LEGEND
    ws.merge_cells("A2:P2")

    niklas = derive_niklas(paul, dom, tm_plan=tm_plan)
    end = add_schedule(ws, 4,
                       "Wochenplan (07:30–18:00) – Paul / Dom / Niklas-Kontrolle",
                       paul, dom, niklas)

    if tm_plan is not None:
        tm_h = tm_hours_from_plan(tm_plan)
        kosten, grp, hh = tm_kosten(96000 * paul_h / 40 + 84000 * dom_h / 40, tm_h)
        end = add_income_table(ws, end + 2, paul_h, dom_h,
                               tm_kosten_monat=kosten, tm_label=f"Gruppe {grp}, {tm_label}")
        ws.cell(row=end, column=1,
                value=f"→ TM: {tm_days_count} Tage, {tm_h}h/Woche, "
                      f"Gruppe {grp} (Haushalt -30% = {int(hh):,} €)".replace(",", ".")).font = FONT_LEGEND
        end += 2
    else:
        end = add_income_table(ws, end + 2, paul_h, dom_h)

    write_legend(ws, end + 1)


# ---------- Übersicht ----------
ws = wb.create_sheet("Übersicht")
ws.column_dimensions['A'].width = 4
ws.column_dimensions['B'].width = 50
ws.column_dimensions['C'].width = 38

ws['B2'] = "Betreuungsplanung Niklas Tom-Hardy – V0.04"
ws['B2'].font = Font(name="Arial", size=18, bold=True, color="1F4E78")

ws['B3'] = ("Update V0.04: + Variante V12 (Dom Do HO 8h, TM 24h). "
            "V0.03: V10/V11 (Paul 35h auf 4 Tage). Zeitraster auf 18:00.")
ws['B3'].font = Font(name="Arial", size=10, italic=True, color="666666")

ws['B5'] = "Annahmen & Rahmenbedingungen"
ws['B5'].font = FONT_SUBHEAD; ws['B5'].fill = COL_SUBHEAD

facts = [
    ("Paul – Vollzeit (40h/Woche)", "96.000 € / Jahr brutto"),
    ("Dominique – Vollzeit (40h/Woche)", "84.000 € / Jahr brutto"),
    ("Paul – Anfahrt", "30 Min. mit Auto"),
    ("Dominique – Anfahrt", "15 Min. zu Fuß"),
    ("Tagesmutter – Standard-Öffnungszeit", "07:30 – 13:00 (5,5h/Tag); abweichend in V10/V11"),
    ("Bringen", "07:30 – Paul mit Auto, Dom. zu Fuß"),
    ("Plan-Raster", "07:30 – 18:00 in 30-Min-Slots"),
    ("Pausenregel", ">6h Tag = 0,5h unbezahlte Pause"),
    ("Homeoffice-Regel", "Während HO ≈ 50% Eigen-Betreuung"),
    ("Niklas – Alter", "unter 3 Jahre (Anlage 1a 'unter 3 J.')"),
    ("Steuerklassen (Netto)", "IV/IV, verheiratet, 1 Kind, Schätzung 2026"),
]
for i, (k, v) in enumerate(facts):
    ws.cell(row=6 + i, column=2, value=k).font = FONT_BOLD
    ws.cell(row=6 + i, column=2).border = BORDER
    ws.cell(row=6 + i, column=3, value=v).font = FONT_NORMAL
    ws.cell(row=6 + i, column=3).border = BORDER

start = 6 + len(facts) + 2
ws.cell(row=start, column=2, value="Ziel").font = FONT_SUBHEAD
ws.cell(row=start, column=2).fill = COL_SUBHEAD
ws.cell(row=start + 1, column=2,
        value="Möglichst viel Zeit mit Niklas, hohes Einkommen, gute Work-Life-Balance.").font = FONT_NORMAL
ws.merge_cells(start_row=start + 1, start_column=2, end_row=start + 1, end_column=4)

start += 4
ws.cell(row=start, column=2, value="Inhalte / Blätter").font = FONT_SUBHEAD
ws.cell(row=start, column=2).fill = COL_SUBHEAD
sheets_overview = [
    ("V1 – Beide Vollzeit (40h)", "Ohne TM."),
    ("V2 – Paul 40h + Dom 50% (20h)", "Ohne TM."),
    ("V3 – Beide 35h", "Ohne TM."),
    ("V4 – Paul 28h (4 Tage) + Dom 40h", "Ohne TM. Paul Fr KOMPLETT FREI."),
    ("V5 – Paul 21h (3 Tage) + Dom 40h", "Ohne TM. Paul Do/Fr FREI."),
    ("V6 – TM 5d + Beide 40h", "TM Mo-Fr 07:30-13:00."),
    ("V7 – TM 5d + Beide 35h", "TM Mo-Fr."),
    ("V8 – TM 3d + Paul 28h + Dom 50%", "TM Mo/Mi/Fr."),
    ("V9 – TM 2d + Paul 21h + Dom 50%", "TM Mo/Mi."),
    ("V10 – TM Mo/Di/Do (20h) + Paul 35h(4d) + Dom 20h",
     "Paul: Mo/Di HO 9h, Mi Office 9h, Do Office 8h, Fr frei. Dom: Mo/Di Office 8h, Do HO 4h."),
    ("V11 – TM Mo/Di (16h) + Paul 35h(4d) + Dom 20h",
     "Wie V10, aber Tagesmutter nur Mo/Di."),
    ("V12 – TM Mo/Di/Do je 8h (24h) + Paul 35h(4d) + Dom 24h",
     "Dom Do HO 8h statt 4h. TM 24h. Höchstes Einkommen unter den 4-Tage-Modellen."),
    ("Vergleich", "Alle Varianten: Brutto, Netto, TM-Kosten, Net-Net."),
]
for i, (k, v) in enumerate(sheets_overview):
    ws.cell(row=start + 1 + i, column=2, value=k).font = FONT_BOLD
    ws.cell(row=start + 1 + i, column=2).border = BORDER
    ws.cell(row=start + 1 + i, column=3, value=v).font = FONT_NORMAL
    ws.cell(row=start + 1 + i, column=3).border = BORDER

start += len(sheets_overview) + 3
write_legend(ws, start)


# ---------- Variant Sheets ----------
build_variant_sheet(
    "V1 Beide Vollzeit",
    "V1 – Beide Vollzeit 40h, ohne Tagesmutter",
    "Mo/Di Paul Office + Dom HO; Mi Hybrid (gespiegelt); Do/Fr Paul HO + Dom Office.",
    *make_v1(), 40, 40)

build_variant_sheet(
    "V2 Dom 50%",
    "V2 – Paul 40h + Dominique 50% (20h), ohne Tagesmutter",
    "Dom 5×4h: Mo/Di/Mi HO (Paul Office), Do/Fr Office (Paul HO).",
    *make_v2(), 40, 20)

build_variant_sheet(
    "V3 Beide 35h",
    "V3 – Beide 35h, ohne Tagesmutter",
    "5×7h pro Person, gegenläufige HO/Office-Verteilung. Mi Hybrid.",
    *make_v3(), 35, 35)

build_variant_sheet(
    "V4 Paul 28h",
    "V4 – Paul 28h (4 Tage) + Dominique 40h, ohne Tagesmutter",
    "Paul: Mo/Di Büro 7h, Mi Hybrid, Do HO 7h. Fr KOMPLETT FREI.",
    *make_v4(), 28, 40)

build_variant_sheet(
    "V5 Paul 21h",
    "V5 – Paul 21h (3 Tage) + Dominique 40h, ohne Tagesmutter",
    "Paul: Mo/Di Büro 7h, Mi HO 7h. Do/Fr komplett FREI.",
    *make_v5(), 21, 40)

build_variant_sheet(
    "V6 TM 5d Vollzeit",
    "V6 – Tagesmutter Mo-Fr + Beide Vollzeit 40h",
    "TM 07:30-13:00 (5×5,5=27,5h, Bracket 25-29h).",
    *make_v6(), 40, 40,
    tm_plan=make_standard_tm_plan(DAYS), tm_label="25-29h", tm_days_count=5)

build_variant_sheet(
    "V7 TM 5d Beide 35h",
    "V7 – Tagesmutter Mo-Fr + Beide 35h",
    "TM 5×5,5h, Eltern je 35h.",
    *make_v7(), 35, 35,
    tm_plan=make_standard_tm_plan(DAYS), tm_label="25-29h", tm_days_count=5)

build_variant_sheet(
    "V8 TM 3d Paul28 Dom50",
    "V8 – Tagesmutter 3 Tage (Mo/Mi/Fr) + Paul 28h + Dominique 50% (20h)",
    "Paul: Mo/Mi Office 7h, Do Hybrid 7h, Fr HO 7h, Di FREI. Dom: 5×4h.",
    *make_v8(), 28, 20,
    tm_plan=make_standard_tm_plan(["Montag", "Mittwoch", "Freitag"]),
    tm_label="15-19h", tm_days_count=3)

build_variant_sheet(
    "V9 TM 2d Paul21 Dom50",
    "V9 – Tagesmutter 2 Tage (Mo/Mi) + Paul 21h + Dominique 50%",
    "Paul: Mo/Mi Office 7h, Fr HO 7h, Di/Do FREI. Dom: 5×4h.",
    *make_v9(), 21, 20,
    tm_plan=make_standard_tm_plan(["Montag", "Mittwoch"]),
    tm_label="10-14h", tm_days_count=2)

# V10 / V11
build_variant_sheet(
    "V10 TM 20h Paul35-4d",
    "V10 – Tagesmutter Mo/Di/Do (20h) + Paul 35h auf 4 Tage + Dominique 20h",
    "Paul: Mo HO 9h, Di HO 9h, Mi Office 9h, Do Office 8h, Fr FREI. "
    "Dom: Mo/Di Office 8h, Mi FREI, Do HO 4h, Fr FREI. "
    "TM: Mo 07:30-15:30 (8h), Di 07:30-15:30 (8h), Do 07:30-11:30 (4h).",
    *make_v10_v11_paul_dom(), 35, 20,
    tm_plan=make_v10_tm(), tm_label="20-24h", tm_days_count=3)

build_variant_sheet(
    "V11 TM 16h Paul35-4d",
    "V11 – Tagesmutter Mo/Di (16h) + Paul 35h auf 4 Tage + Dominique 20h",
    "Paul/Dom wie V10. TM nur Mo/Di je 8h = 16h. "
    "Donnerstag: keine TM, Paul Office, Dom HO 4h – Dom betreut.",
    *make_v10_v11_paul_dom(), 35, 20,
    tm_plan=make_v11_tm(), tm_label="15-19h", tm_days_count=2)

build_variant_sheet(
    "V12 TM 24h Paul35-4d",
    "V12 – Tagesmutter Mo/Di/Do je 8h (24h) + Paul 35h auf 4 Tage + Dominique 24h",
    "Paul wie V10. Dom: Mo/Di Office 8h, Do HO 8h, Mi/Fr FREI = 24h. "
    "TM: Mo/Di/Do je 07:30-15:30 (8h) = 24h.",
    *make_v12(), 35, 24,
    tm_plan=make_v12_tm(), tm_label="20-24h", tm_days_count=3)


# ---------- Vergleichs-Sheet ----------
ws = wb.create_sheet("Vergleich")
ws.column_dimensions['A'].width = 32
for col in range(2, 14):
    ws.column_dimensions[get_column_letter(col)].width = 13

ws.cell(row=1, column=1, value="Variantenvergleich").font = FONT_HEADER
ws.cell(row=1, column=1).fill = COL_HEADER
ws.merge_cells("A1:L1")
ws.cell(row=1, column=1).alignment = CENTER

headers = ["Variante", "Paul Std/W", "Dom Std/W",
           "Paul Brutto/Mon", "Dom Brutto/Mon", "Brutto HH/Mon",
           "Paul Netto/Mon", "Dom Netto/Mon", "Netto HH/Mon",
           "TM/Mon", "Netto-Net/Mon", "Eltern-Zeit*"]
for i, h in enumerate(headers):
    c = ws.cell(row=3, column=1 + i, value=h)
    c.font = FONT_BOLD; c.fill = COL_TOTAL; c.border = BORDER; c.alignment = CENTER

variants_data = [
    ("V1 Beide Vollzeit", 40, 40, None, "niedrig"),
    ("V2 Paul 40h + Dom 50%", 40, 20, None, "mittel"),
    ("V3 Beide 35h", 35, 35, None, "mittel-hoch"),
    ("V4 Paul 28h + Dom 40h", 28, 40, None, "hoch"),
    ("V5 Paul 21h + Dom 40h", 21, 40, None, "sehr hoch"),
    ("V6 TM 5d + Beide 40h", 40, 40, "v6", "niedrig (TM stabil)"),
    ("V7 TM 5d + Beide 35h", 35, 35, "v7", "mittel"),
    ("V8 TM 3d + Paul 28h + Dom 50%", 28, 20, "v8", "sehr hoch"),
    ("V9 TM 2d + Paul 21h + Dom 50%", 21, 20, "v9", "maximal"),
    ("V10 TM 20h + Paul 35h(4d) + Dom 20h", 35, 20, "v10", "hoch (Paul Fr frei)"),
    ("V11 TM 16h + Paul 35h(4d) + Dom 20h", 35, 20, "v11", "hoch (Paul Fr frei)"),
    ("V12 TM 24h + Paul 35h(4d) + Dom 24h", 35, 24, "v12", "hoch (Paul Fr frei)"),
]
tm_lookup = {
    "v6": tm_kosten(96000 + 84000, 27.5)[0],
    "v7": tm_kosten(96000 * 35 / 40 + 84000 * 35 / 40, 27.5)[0],
    "v8": tm_kosten(96000 * 28 / 40 + 84000 * 20 / 40, 16.5)[0],
    "v9": tm_kosten(96000 * 21 / 40 + 84000 * 20 / 40, 11)[0],
    "v10": tm_kosten(96000 * 35 / 40 + 84000 * 20 / 40, 20)[0],
    "v11": tm_kosten(96000 * 35 / 40 + 84000 * 20 / 40, 16)[0],
    "v12": tm_kosten(96000 * 35 / 40 + 84000 * 24 / 40, 24)[0],
}

for i, (name, ph, dh, tm_key, note) in enumerate(variants_data):
    r = 4 + i
    ws.cell(row=r, column=1, value=name).font = FONT_BOLD
    ws.cell(row=r, column=2, value=ph)
    ws.cell(row=r, column=3, value=dh)
    ws.cell(row=r, column=4, value=f"=B{r}/40*96000/12")
    ws.cell(row=r, column=5, value=f"=C{r}/40*84000/12")
    ws.cell(row=r, column=6, value=f"=D{r}+E{r}")
    ws.cell(row=r, column=7, value=_netto_formula(f"D{r}"))
    ws.cell(row=r, column=8, value=_netto_formula(f"E{r}"))
    ws.cell(row=r, column=9, value=f"=G{r}+H{r}")
    tm_val = tm_lookup.get(tm_key, 0) if tm_key else 0
    ws.cell(row=r, column=10, value=tm_val)
    ws.cell(row=r, column=11, value=f"=I{r}-J{r}")
    ws.cell(row=r, column=12, value=note)
    for col in range(1, 13):
        ws.cell(row=r, column=col).border = BORDER
        ws.cell(row=r, column=col).alignment = CENTER
    for col in (4, 5, 6, 7, 8, 9, 10, 11):
        ws.cell(row=r, column=col).number_format = '#,##0 "€"'
    ws.cell(row=r, column=11).fill = COL_INCOME
    ws.cell(row=r, column=11).font = FONT_BOLD

note_row = 4 + len(variants_data) + 2
notes = [
    "* Eltern-Zeit qualitativ: verfügbare Aufmerksamkeit für Niklas im Wochenmittel.",
    "Während HO ist Eigen-Betreuung nur zu ca. 50 % möglich.",
    "TM-Kosten gem. Anlage 1a Landkreis Tübingen, gültig 01.01.2026 – 31.08.2026, 'unter 3 J.'.",
    "Einkommensgruppe = (Brutto Haushalt − 30 %) gem. Satzung § 3 (2).",
    "Netto: Schätzung Steuerklasse IV/IV, verheiratet, 1 Kind, 2026 (brutto-netto-rechner.info).",
    "Netto-Faktoren je Brutto/Mon: ≥8000=60%, ≥7000=61%, ≥6000=62%, ≥5000=64%, ≥4000=66%, ≥3000=68%, sonst 70%.",
    "V10/V11: TM-Zeitfenster abweichend von Standard 07:30-13:00 (siehe jeweiliges Blatt).",
]
for i, n in enumerate(notes):
    ws.cell(row=note_row + i, column=1, value=n).font = FONT_LEGEND
    ws.merge_cells(start_row=note_row + i, start_column=1,
                   end_row=note_row + i, end_column=12)


wb.move_sheet("Übersicht", offset=-len(wb.sheetnames))
wb.save(OUT)
print(f"OK -> {OUT}")
